"""Post-process a downloaded GL Account Detail xlsx per user's manual workflow.

Steps 1-5 from the reconciliation spec:
  1. Remove header rows (title / date range / legal entity / blanks)
  2. Remove columns A & B (leading empty columns)
  3. Unmerge all cells + apply wrap-text to every cell
  4. Insert an "Amount" column immediately after Credit
  5. Populate Amount with =Debit-Credit (Debit positive, Credit negative)

Steps 7+ (color-coded reconciliation and JE posting) are not automated yet.
"""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

log = logging.getLogger(__name__)

WRAP = Alignment(wrap_text=True, vertical="top")

COLUMN_WIDTHS = {
    "A": 22.0,
    "B": 11.0,
    "C": 14.0,
    "D": 16.0,
    "E": 14.0,
    "F": 18.0,
    "G": 24.0,
    "H": 11.0,
    "I": 11.0,
    "J": 11.0,
    "K": 13.0,
}


def _find_header_row(ws) -> tuple[int, int]:
    """Locate the header row and the 1-based column index of 'Date'.

    Returns (row_number, date_col_index).
    """
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25)):
        values = [str(c.value).strip().lower() if c.value else "" for c in row]
        if "date" in values and "debit" in values and "credit" in values:
            date_col = next(c.column for c in row if c.value and str(c.value).strip().lower() == "date")
            return row[0].row, date_col
    raise ValueError("Could not find header row (Date/Debit/Credit) in sheet")


def _unmerge_all(ws) -> None:
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))


def process_gl_workbook(src: str | Path, dest: str | Path | None = None) -> Path:
    """Read the downloaded GL xlsx, apply steps 1-5, write processed xlsx.

    If dest is None, writes alongside src with `_processed` suffix.
    Returns the path to the written file.
    """
    src = Path(src)
    if dest is None:
        dest = src.with_name(src.stem + "_processed.xlsx")
    else:
        dest = Path(dest)

    wb = load_workbook(src)
    ws = wb.active

    _unmerge_all(ws)

    header_row, date_col = _find_header_row(ws)
    log.info("Header row detected at row %d, Date at col %d", header_row, date_col)

    if header_row > 1:
        ws.delete_rows(1, header_row - 1)

    leading_empty = date_col - 1
    if leading_empty > 0:
        ws.delete_cols(1, leading_empty)
        log.info("Removed %d leading empty column(s)", leading_empty)

    header_cells = {
        (str(c.value).strip().lower() if c.value else ""): c.column
        for c in ws[1]
    }
    credit_col = header_cells.get("credit")
    if not credit_col:
        raise ValueError("Credit column not found after cleanup")

    amount_col = credit_col + 1
    ws.insert_cols(amount_col)
    ws.cell(row=1, column=amount_col, value="Amount")

    debit_col = header_cells.get("debit")
    if not debit_col:
        raise ValueError("Debit column not found after cleanup")

    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=debit_col).value
        c = ws.cell(row=r, column=credit_col).value
        if isinstance(d, (int, float)) or isinstance(c, (int, float)):
            d_val = d if isinstance(d, (int, float)) else 0
            c_val = c if isinstance(c, (int, float)) else 0
            ws.cell(row=r, column=amount_col, value=round(d_val - c_val, 2))

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP

    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    wb.save(dest)
    log.info("Processed GL saved: %s", dest)
    return dest


if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    if len(sys.argv) < 2:
        print("Usage: python -m r365.gl_excel_processor <input.xlsx> [output.xlsx]")
        sys.exit(1)
    out = process_gl_workbook(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
    print(out)
